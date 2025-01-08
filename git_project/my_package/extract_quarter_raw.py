import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side
import sys
import time

os.system('cls' if os.name == 'nt' else 'clear')  # Clear screen (Windows: cls, Linux/Mac: clear)

#Finding source file thru relative path
def find_source_file(file, dir):
    for root, dirs, files in os.walk(dir):
        
        path_parts = root.split(os.sep)  # Split the path into parts
        
        # Display the first level, second level, and final level
        if len(path_parts) >= 3:
            display_path = f"{path_parts[0]}\\{path_parts[1]}\\...\\{path_parts[-1]}"
        elif len(path_parts) == 2:
            display_path = f"{path_parts[0]}\\{path_parts[1]}"
        else:
            display_path = root
        
        sys.stdout.write(f"\rFinding {file} in {display_path}")  # Write without a newline
        sys.stdout.flush()  # Flush the buffer to update the screen immediately
        time.sleep(0.1)
        if file in files:
            print("\nFile Found!\n")
            return os.path.join(root, file)
    else:
        print(f"\n\n{file} Not Found!")
        print("Terminating the program")
        return sys.exit(1)

#Finding output file thru relative path
def find_output_file(save, dir):
    for root, dirs, files in os.walk(dir):
        
        path_parts = root.split(os.sep)  # Split the path into parts
        
        # Display the first level, second level, and final level
        if len(path_parts) >= 3:
            display_path = f"{path_parts[0]}\\{path_parts[1]}\\...\\{path_parts[-1]}"
        elif len(path_parts) == 2:
            display_path = f"{path_parts[0]}\\{path_parts[1]}"
        else:
            display_path = root
            
        sys.stdout.write(f"\rFinding {save} in {display_path}")
        sys.stdout.flush()  # Flush the buffer to update the screen immediately
        time.sleep(0.1)
        if save in files:
            print("\nFile Found!\n")
            return os.path.join(root, save)
    else:
        print(f"\n\n{save} Not Found!")
        print("Terminating the program")
        return sys.exit(1)  

def extract_quarter_raw(dir, file, save, sheets):
    """
    Extracts raw data for quarterly report

    Args:
        dir(str): working directory of file
        file(str): file name to extract data from
        save(str): save file name to extract to
        sheets(str): month sheet to include during extraction. ex: ["sheet1", "sheet2", "sheet(n+1)"]
    """
    try:
        #Declaring working directory
        working_directory = dir

        # Excel file path and save path
        file_init = find_source_file(file, dir)
        save_path = find_output_file(save, dir)
        
        sys.stdout.write(f"Extracting Report...\n")  # Write without a newline
        sys.stdout.flush()  # Flush the buffer to update the screen immediately
        time.sleep(0.5)

        # Changing the working directory
        os.chdir(working_directory)

        # Check if file exists, if not, create a new workbook
        if not os.path.exists(save_path):
            wb = Workbook()  # Create a new Workbook
            ws = wb.active  # Get the active worksheet
            ws.title = "Sheet1"  # Optionally, rename the active worksheet
            wb.save(save_path)  # Save the workbook with the desired filename

        # Preloading save path excel file
        excel_file = pd.ExcelFile(save_path)
        #print(excel_file.sheet_names)
        # Sheets to override and exclude
        sheets_to_override = ["RECEIVED_DATA(DO NOT PRINT)", "TESTED_DATA(DO NOT PRINT)"]
        sheets_to_exclude = [sheet for sheet in excel_file.sheet_names if sheet not in sheets_to_override]

        # Load the workbook
        wb = load_workbook(save_path)

        # Remove specific sheets from the existing workbook
        for sheet_name in sheets_to_override:
            if sheet_name in wb.sheetnames:
                std = wb[sheet_name]
                wb.remove(std)

        # Save after removing sheets
        wb.save(save_path)

        # Reading data from specified sheets and concatenating them
        column_names = ["DATE_RECEIVED", "RECEIVED_BY", "SECTOR", "COMPANY", "QTY", "TEST_METHOD", 
                        "DESCRIPTION", "DATE_TESTED", "TESTED_BY", "DETAILS", "PASSED", "FAILED", 
                        "RELEASED_DATE", "BRAND"]

        # Reading sheets and appending the results
        df_fin = []
        
        try:
            for sheet in sheets[0:2]:
                df = pd.read_excel(file_init, sheet_name=sheet, skiprows=2, names=column_names)
                #df = df.astype(str)
                df = df[df["DATE_TESTED"].notna()]
                df["DESCRIPTION"] = df["DESCRIPTION"].str.strip()
                df_fin.append(df)
            for sheet in sheets[-1:]:
                df = pd.read_excel(file_init, sheet_name=sheet, skiprows=2, names=column_names)
                #df = df.astype(str)
                df = df[df["DATE_RECEIVED"].notna()]
                df["DESCRIPTION"] = df["DESCRIPTION"].str.strip()
                df_fin.append(df)
        except Exception as e:
            print(f"Unexpected Error: {e}")
            #print("\nSuggested Solution: Check the list of Worksheets to extract data from.\n")
            sys.exit(1)

        # Concatenating all dataframes
        df_fin = pd.concat(df_fin, ignore_index=True)

        # Convert 'DATE_TESTED' to datetime format
        df_fin['DATE_TESTED'] = pd.to_datetime(df_fin['DATE_TESTED'], errors='coerce')
        df_fin['DATE_RECEIVED'] = pd.to_datetime(df_fin['DATE_RECEIVED'], errors='coerce')
        df_fin['QTY'] = df_fin['QTY'].astype(float)
        df_fin['PASSED'] = df_fin['PASSED'].astype(float)
        df_fin['FAILED'] = df_fin['FAILED'].astype(float)
        print(df_fin[["DESCRIPTION", "DATE_TESTED", "PASSED", "FAILED"]])
        print(df_fin.isna().sum())
        print(df.info())

        # Split data into two sets
        #df_fin_2= df_fin[df_fin["DATE_TESTED"].notna()]
        inquiry_received = df_fin[["DATE_RECEIVED", "RECEIVED_BY", "SECTOR", "COMPANY", "QTY", "TEST_METHOD", "DESCRIPTION", "DETAILS"]]
        #inquiry_received = inquiry_received.drop_duplicates(subset="DETAILS", keep='first')
        inquiry_tested = df_fin[["DATE_RECEIVED", "RECEIVED_BY", "SECTOR", "COMPANY", "QTY", "TEST_METHOD", 
                                "DESCRIPTION", "DATE_TESTED", "TESTED_BY", "DETAILS", "PASSED", "FAILED"]]

        # Save to the designated sheets in one ExcelWriter block
        with pd.ExcelWriter(save_path, mode='a', engine='openpyxl') as writer:
            inquiry_received.to_excel(writer, sheet_name="RECEIVED_DATA(DO NOT PRINT)", index=False)
            inquiry_tested.to_excel(writer, sheet_name="TESTED_DATA(DO NOT PRINT)", index=False)

        # Reload workbook to apply styling
        wb = load_workbook(save_path)

        # Apply formatting (borders, tab color)
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for sheet in sheets_to_override:
            if sheet in wb.sheetnames:
                ws = wb[sheet]
                ws.sheet_properties.tabColor = "DDEF31"  # Set the tab color

                # Apply borders to all cells in the sheet
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border

        # Save the workbook with applied changes
        wb.save(save_path)
        print("Quarterly Report Successfully Extracted!")
    except PermissionError as e:
        print(f"\nError: {e}")
        #print("\nSuggested Solution: Check if you have permission to access the directory or the file.\n")
        sys.exit(1)  # Exit the program if the file is not found
    except Exception as e:
        print(f"\nUnexpected Error: {e}")
        #print("\nSuggested Solution: Ensure that the directory path is correct and the system has access to it.\n")
        sys.exit(1)